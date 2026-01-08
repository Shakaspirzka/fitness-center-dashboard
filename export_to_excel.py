"""
Script pentru exportarea datelor în Excel
"""
import pandas as pd
from calculations import compare_scenarios, OCCUPANCY_SCENARIOS, SUBSCRIPTION_TYPES
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def export_analysis_to_excel(
    subscription_distribution: dict,
    participation_rate: float = 0.10,
    population_density: float = 1000,
    filename: str = "analiza_fitness_center.xlsx"
):
    """
    Exportă analiza completă în Excel
    """
    wb = Workbook()
    
    # Sheet 1: Comparare Scenarii
    ws1 = wb.active
    ws1.title = "Comparare Scenarii"
    
    comparison_df = compare_scenarios(
        subscription_distribution,
        participation_rate,
        population_density
    )
    
    # Adaugă header
    ws1.append(['Comparare Scenarii de Ocupare'])
    ws1.append([])
    
    # Adaugă datele
    for r in dataframe_to_rows(comparison_df, index=False, header=True):
        ws1.append(r)
    
    # Formatare header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws1[3]:  # Rândul cu header-ul
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 2: Detalii Abonamente
    ws2 = wb.create_sheet("Detalii Abonamente")
    
    ws2.append(['Tipuri de Abonamente Disponibile'])
    ws2.append([])
    ws2.append(['Tip', 'Preț (RON)', 'Sesiuni', 'Procentaj Distribuție'])
    
    for key, sub_type in SUBSCRIPTION_TYPES.items():
        pct = subscription_distribution.get(key, 0) * 100
        sessions = sub_type['sessions'] if sub_type['sessions'] else 'Nelimitat'
        ws2.append([sub_type['name'], sub_type['price'], sessions, f"{pct:.1f}%"])
    
    # Formatare
    for cell in ws2[3]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Sheet 3: Parametri
    ws3 = wb.create_sheet("Parametri")
    
    ws3.append(['Parametri Analiză'])
    ws3.append([])
    ws3.append(['Parametru', 'Valoare'])
    ws3.append(['Rata Participare Populație', f"{participation_rate*100:.1f}%"])
    ws3.append(['Densitate Populație', f"{population_density:,} oameni/km²"])
    ws3.append(['Capacitate per Oră', '20 oameni'])
    ws3.append(['Ore pe Zi', '10 ore'])
    ws3.append(['Zile pe Săptămână', '7 zile'])
    ws3.append(['Venit Dorit Lunar', '50,000 RON'])
    
    # Formatare
    for cell in ws3[3]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Salvează
    wb.save(filename)
    print(f"Fișier Excel creat: {filename}")

if __name__ == "__main__":
    # Exemplu de utilizare
    subscription_dist = {
        'economic': 0.4,
        'standard': 0.5,
        'premium': 0.1
    }
    
    export_analysis_to_excel(
        subscription_dist,
        participation_rate=0.10,
        population_density=1000
    )

